import os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from exporters.base_exporter import BaseExporter
from core.models import RawProduct

class ExcelExporter(BaseExporter):
    def export(self, products: List[RawProduct], filename: str = "") -> str:
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"scraped_products_{timestamp}.xlsx"
        
        output_dir = "output"
        os.makedirs(output_dir, exist_ok=True)
        filepath = os.path.join(output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # 1. Collect all unique spec keys from all products
        all_spec_keys = []
        for p in products:
            specs = getattr(p, 'raw_specs', {}) or {}
            for key in specs.keys():
                if key not in all_spec_keys and key != "is_relevant":
                    all_spec_keys.append(key)
        
        # Sort keys to have Brand/Model first if they exist
        priority_keys = ["Category", "Brand", "Model", "Voltage", "Capacity"]
        sorted_keys = [k for k in priority_keys if k in all_spec_keys]
        sorted_keys += [k for k in all_spec_keys if k not in priority_keys]

        # 2. Define Headers
        base_headers = ["Marketplace", "Title", "Price (UAH)", "Is Relevant"]
        final_headers = base_headers + sorted_keys + ["URL", "Found At"]
        ws.append(final_headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # 3. Fill Rows
        for p in products:
            row_data = []
            # Base info
            row_data.append(p.marketplace)
            row_data.append(p.title)
            row_data.append(p.price)
            
            specs = getattr(p, 'raw_specs', {}) or {}
            is_rel = "YES" if specs.get("is_relevant", True) else "NO"
            row_data.append(is_rel)
            
            # Dynamic specs
            for key in sorted_keys:
                row_data.append(specs.get(key, "-"))
                
            # Tail info
            row_data.append(p.url)
            row_data.append(p.scraped_at.strftime("%Y-%m-%d %H:%M:%S") if p.scraped_at else "N/A")
            
            ws.append(row_data)

        # 4. Adjust column widths
        for i, column_cells in enumerate(ws.columns):
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            # Constraints
            adj_width = min(max_length + 2, 60)
            ws.column_dimensions[column].width = adj_width

        wb.save(filepath)
        return filepath
