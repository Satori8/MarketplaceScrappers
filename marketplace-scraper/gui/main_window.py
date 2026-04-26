import tkinter as tk
import customtkinter as ctk
from tkinter import messagebox, simpledialog, filedialog, ttk
import queue
import threading
import uuid
import logging
import concurrent.futures
import asyncio
from datetime import datetime, timezone
from typing import List, Dict

from gui.styles import COLORS, FONTS, apply_styles
from core.models import ScrapeTask, RawProduct
from core.scheduler import TaskScheduler
from db.database import Database

logger = logging.getLogger(__name__)

class MainWindow(ctk.CTkFrame):
    def __init__(self, master, db_path="data/products.db"):
        super().__init__(master)
        self.master = master
        self.db = Database(db_path)
        self.db.initialize()
        self.scheduler = TaskScheduler(self.db, on_keys_exhausted=self._on_gemini_keys_exhausted)
        self.scheduler.on_mp_status = lambda mp, status: self.msg_queue.put(("mp_status", mp, status))
        
        self.msg_queue = queue.Queue()
        self.is_running = False
        self.total_queries = 0
        self.queries_completed = 0
        
        self.marketplaces_vars = {}
        self.marketplaces_methods = {}
        self.marketplaces_status_labels = {}
        self.all_found: List[RawProduct] = []
        self.log_widgets: Dict[str, ctk.CTkTextbox] = {}
        
        # Target Links Storage
        self.target_links = []
        
        self._setup_ui()
        self._setup_logging()
        self._setup_callbacks()
        self._poll_queue()

    def _setup_logging(self):
        class MultiChannelHandler(logging.Handler):
            def __init__(self, q):
                super().__init__()
                self.queue = q
            def emit(self, record):
                group = "SYSTEM"
                name = record.name.lower()
                
                if "gemini" in name or "ai." in name: group = "AI"
                elif "scheduler" in name: group = "SCHEDULER"
                elif "scrapers." in name or "scraper" in name:
                    parts = record.name.split(".")
                    group = parts[1].upper() if len(parts) > 1 else "MARKETPLACES"
                
                self.queue.put(("log", group, record.levelname, self.format(record)))
        
        handler = MultiChannelHandler(self.msg_queue)
        handler.setFormatter(logging.Formatter('%(asctime)s - %(message)s', '%H:%M:%S'))
        root_logger = logging.getLogger()
        root_logger.addHandler(handler)
        root_logger.setLevel(logging.INFO)

    def _setup_ui(self):
        self.master.title("Marketplace Scraper Intelligence v1.5")
        self.master.geometry("1400x900")
        apply_styles()

        # Layout
        self.sidebar = ctk.CTkFrame(self.master, width=280, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        
        self.content = ctk.CTkFrame(self.master, corner_radius=0, fg_color="transparent")
        self.content.pack(side="right", expand=True, fill="both", padx=20, pady=20)

        # sidebar components
        main_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        ctk.CTkLabel(main_frame, text="AGENT SETTINGS", font=FONTS["title"], text_color=COLORS["accent"]).pack(pady=(0, 20))
        
        ctk.CTkLabel(main_frame, text="Search Keywords:").pack(anchor="w")
        self.query_entry = ctk.CTkEntry(main_frame, placeholder_text="e.g. lifepo4 100ah 12v")
        self.query_entry.pack(fill="x", pady=(0, 15))
        self.query_entry.insert(0, "lifepo4 100ah 12v, lifepo4 100ah 24v")

        ctk.CTkLabel(main_frame, text="Pages:").pack(anchor="w")
        self.pages_var = tk.StringVar(value="1")
        self.pages_entry = ctk.CTkEntry(main_frame, textvariable=self.pages_var, width=60)
        self.pages_entry.pack(pady=(0, 15), anchor="w")

        link_ctrl = ctk.CTkFrame(main_frame, fg_color="transparent")
        link_ctrl.pack(fill="x", pady=(0, 20))
        self.use_links_var = tk.BooleanVar(value=False)
        ctk.CTkCheckBox(link_ctrl, text="Use Target Links", variable=self.use_links_var).pack(side="left")
        ctk.CTkButton(link_ctrl, text="LINKS", command=self._on_link_entry_click, width=80, height=28).pack(side="right")

        ctk.CTkLabel(main_frame, text="Marketplaces:", font=FONTS["title"], text_color=COLORS["accent"]).pack(anchor="w", pady=(0, 10))
        mp_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        mp_frame.pack(fill="x")
        
        self.active_mps = ["HOTLINE", "ROZETKA", "PROM", "ALLO", "EPICENTRK"]
        for mp in self.active_mps:
            m_low = mp.lower()
            var = tk.BooleanVar(value=True)
            self.marketplaces_vars[m_low] = var
            row = ctk.CTkFrame(mp_frame, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkCheckBox(row, text=mp, variable=var).pack(side="left")
            mc = ctk.CTkComboBox(row, values=["Auto", "Requests", "Browser"], width=100, height=28)
            mc.set("Browser" if m_low in ["rozetka", "hotline"] else "Auto")
            mc.pack(side="right")
            self.marketplaces_methods[m_low] = mc
            sl = ctk.CTkLabel(row, text="●", text_color="gray", font=("Arial", 16))
            sl.pack(side="right", padx=10)
            self.marketplaces_status_labels[m_low] = sl

        self.start_btn = ctk.CTkButton(main_frame, text="LAUNCH AGENTS", command=self._on_start_click, height=45, font=FONTS["title"])
        self.start_btn.pack(pady=(30, 10), fill="x")
        self.stop_btn = ctk.CTkButton(main_frame, text="ABORT TASK", state="disabled", command=self._on_stop_click, height=45, fg_color=COLORS["error"], hover_color="#c0392b")
        self.stop_btn.pack(pady=0, fill="x")

        ctk.CTkLabel(main_frame, text="").pack(pady=10)
        ctk.CTkButton(main_frame, text="Database Browser", command=self._on_db_browser_click, fg_color=COLORS["sidebar"]).pack(fill="x", pady=5)
        ctk.CTkButton(main_frame, text="Export Excel", command=self._on_excel_click, fg_color=COLORS["sidebar"]).pack(fill="x", pady=5)
        ctk.CTkButton(main_frame, text="Normalize File", command=self._on_normalize_excel_click, fg_color=COLORS["sidebar"]).pack(fill="x", pady=5)

        # Content
        self.summary_frame = ctk.CTkFrame(self.content, fg_color="transparent")
        self.summary_frame.pack(fill="x", pady=(0, 15))
        self.status_label = ctk.CTkLabel(self.summary_frame, text="System Ready", font=FONTS["title"])
        self.status_label.pack(side="left")
        self.batch_label = ctk.CTkLabel(self.summary_frame, text="", font=FONTS["small"], text_color=COLORS["accent"])
        self.batch_label.pack(side="right")
        
        self.progress = ctk.CTkProgressBar(self.content, mode="determinate")
        self.progress.pack(fill="x", pady=(0, 20))
        self.progress.set(0)

        self.log_tabs = ctk.CTkTabview(self.content, height=500)
        self.log_tabs.pack(fill="both", expand=True)
        
        self.log_tabs.add("LIVE DATA")
        data_frame = self.log_tabs.tab("LIVE DATA")
        cols = ("MP", "Title", "Brand", "Model", "V", "Ah", "Price")
        self.tree = ttk.Treeview(data_frame, columns=cols, show="headings", height=15)
        widths = {"MP": 80, "Title": 400, "Brand": 100, "Model": 100, "V": 50, "Ah": 50, "Price": 80}
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=widths.get(c, 100), anchor="center")
        
        self.tree.pack(side="left", fill="both", expand=True)
        data_scroll = ttk.Scrollbar(data_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=data_scroll.set)
        data_scroll.pack(side="right", fill="y")

        tab_names = ["ALL", "AI", "SCHEDULER"] + self.active_mps + ["SYSTEM"]
        for name in tab_names:
            if name == "LIVE DATA": continue
            self.log_tabs.add(name)
            tab_frame = self.log_tabs.tab(name)
            txt = ctk.CTkTextbox(tab_frame, font=("Consolas", 12), border_width=0)
            txt.pack(fill="both", expand=True)
            txt.tag_config("ERROR", foreground="#ff5555")
            txt.tag_config("WARNING", foreground="#ffb86c")
            txt.tag_config("SUCCESS", foreground="#50fa7b")
            txt.tag_config("INFO", foreground="#f8f8f2")
            txt.configure(state="disabled")
            self.log_widgets[name] = txt

    def _setup_callbacks(self):
        self.scheduler.on_product_found = lambda p, n, d: self.msg_queue.put(("product", p, n, d))
        self.scheduler.on_error = lambda m: self.msg_queue.put(("error", m))
        self.scheduler.on_finished = lambda sid, s: self.msg_queue.put(("query_finished", sid, s))
        self.scheduler.on_progress = lambda c, t: self.msg_queue.put(("progress", c, t))

    def _on_start_click(self):
        queries = [q.strip() for q in self.query_entry.get().split(",") if q.strip()]
        selected_mps = {mp: self.marketplaces_methods[mp].get() for mp, v in self.marketplaces_vars.items() if v.get()}

        if not queries or not selected_mps:
            messagebox.showwarning("Incomplete", "Please enter keywords and select at least one Marketplace.")
            return

        try:
            pages = int(self.pages_var.get())
        except ValueError:
            pages = 1
            
        self.total_queries = len(queries)
        self.queries_completed = 0
        self.all_found = []
        
        # B22 fix: Must clear stop event for subsequent runs
        self.scheduler._stop_event.clear()

        def run_batch():
            self.is_running = True
            self.start_btn.configure(state="disabled")
            self.stop_btn.configure(state="normal")
            
            discovery_tasks = []
            session_id = str(uuid.uuid4())
            for q in queries:
                for mp, method in selected_mps.items():
                    discovery_tasks.append((mp, method, q))
            
            if self.use_links_var.get() and self.target_links:
                for url in self.target_links:
                    mp_detect = None
                    url_l = url.lower()
                    if "rozetka" in url_l: mp_detect = "rozetka"
                    elif "hotline" in url_l: mp_detect = "hotline"
                    elif "prom" in url_l: mp_detect = "prom"
                    elif "allo" in url_l: mp_detect = "allo"
                    elif "epicentr" in url_l: mp_detect = "epicentrk"
                    if mp_detect:
                        meth = selected_mps.get(mp_detect, "Browser" if mp_detect == "rozetka" else "Auto")
                        discovery_tasks.append((mp_detect, meth, url))

            if not discovery_tasks:
                self.msg_queue.put(("batch_finished", None))
                return

            self.msg_queue.put(("status_update", f"Searching across {len(discovery_tasks)} marketplace jobs..."))
            max_workers = min(len(discovery_tasks), 10) 
            with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [
                    executor.submit(self.scheduler.run_individual_discovery, mp, meth, q, pages, session_id)
                    for mp, meth, q in discovery_tasks
                ]
                completed = 0
                for _ in concurrent.futures.as_completed(futures):
                    if not self.is_running: break
                    completed += 1
                    self.queries_completed = (completed * self.total_queries) // len(discovery_tasks)
            
            if self.is_running:
                self.msg_queue.put(("status_update", "AI Intelligence Phase: Mapping Specifications..."))
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    loop.run_until_complete(self.scheduler.normalize_all_pending(stop_event=self.scheduler._stop_event))
                finally:
                    loop.close()
            
            self.msg_queue.put(("batch_finished", None))

        threading.Thread(target=run_batch, daemon=True).start()

    def _on_normalize_excel_click(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="Select Excel File")
        if not file_path: return
        import openpyxl
        import os
        def work():
            try:
                self.after(0, lambda: self.status_label.configure(text="Processing Excel...", text_color=COLORS["accent"]))
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                headers = [str(cell.value).lower().strip() for cell in ws[1]]
                title_idx = next((i for i, h in enumerate(headers) if h in ["title", "name", "product", "назва"]), -1)
                if title_idx == -1:
                    self.after(0, lambda: messagebox.showerror("Error", "Could not find Title column."))
                    return
                raws = []
                for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    title = row_data[title_idx]
                    if not title: continue
                    rp = RawProduct(title=str(title), price=0, marketplace="Import", raw_specs={}, scraped_at=datetime.now(timezone.utc))
                    rp._row_idx = row_idx
                    raws.append(rp)
                if not raws: return
                self.after(0, lambda: self.status_label.configure(text=f"AI Normalizing {len(raws)} items..."))
                normalized = asyncio.run(self.scheduler.normalizer.normalize_batch(raws, query="Excel Import", stop_event=self.scheduler._stop_event))
                new_cols = {"Normalized Brand": len(headers)+1, "Normalized Model": len(headers)+2, "V": len(headers)+3, "Ah": len(headers)+4}
                for label, col in new_cols.items(): ws.cell(row=1, column=col, value=label)
                for norm in normalized:
                    r_idx = getattr(norm.raw, "_row_idx")
                    ws.cell(row=r_idx, column=new_cols["Normalized Brand"], value=norm.normalized_specs.get("brand", ""))
                    ws.cell(row=r_idx, column=new_cols["Normalized Model"], value=norm.normalized_specs.get("model", ""))
                    ws.cell(row=r_idx, column=new_cols["V"], value=norm.normalized_specs.get("voltage", ""))
                    ws.cell(row=r_idx, column=new_cols["Ah"], value=norm.normalized_specs.get("capacity", ""))
                out_path = f"{os.path.splitext(file_path)[0]}_normalized.xlsx"
                wb.save(out_path)
                self.after(0, lambda: self.status_label.configure(text="Excel Normalized", text_color=COLORS["success"]))
                self.after(0, lambda: messagebox.showinfo("Success", f"Saved to {out_path}"))
            except Exception as e:
                logger.error(e)
                self.after(0, lambda: messagebox.showerror("Error", str(e)))
        threading.Thread(target=work, daemon=True).start()

    def _on_stop_click(self):
        self.is_running = False
        self.scheduler.stop()
        self.status_label.configure(text="Aborting Task...", text_color=COLORS["error"])

    def _poll_queue(self):
        if not self.master.winfo_exists(): return
        while not self.msg_queue.empty():
            self._handle_msg(self.msg_queue.get())
        self.master.after(100, self._poll_queue)

    def _handle_msg(self, msg):
        mtype, *args = msg
        if mtype == "log":
            group, level, text = args
            for target in ["ALL", group]:
                if target in self.log_widgets:
                    w = self.log_widgets[target]
                    w.configure(state="normal")
                    w.insert("end", text + "\n", level)
                    w.see("end")
                    w.configure(state="disabled")
        elif mtype == "product":
            prod, is_new, delta = args
            delta_str = f" (Δ{delta:+.0f})" if delta is not None else ""
            self.tree.insert("", "end", values=(prod.marketplace, prod.title, "", "", "", "", f"{prod.price}{delta_str}"))
            self.all_found.append(prod)
        elif mtype == "status_update":
            self.status_label.configure(text=args[0], text_color=COLORS["accent"])
            self.batch_label.configure(text=f"TASK: {self.queries_completed+1}/{self.total_queries}")
        elif mtype == "batch_finished":
            self.is_running = False
            self.start_btn.configure(state="normal")
            self.stop_btn.configure(state="disabled")
            self.status_label.configure(text="Batch Completed Successfully", text_color=COLORS["success"])
            self.progress.set(1.0)
        elif mtype == "mp_status":
            mp, status = args
            if mp.lower() in self.marketplaces_status_labels:
                lbl = self.marketplaces_status_labels[mp.lower()]
                lbl.configure(text="●", text_color=COLORS["success"] if "Finished" in status else COLORS["accent"])
        elif mtype == "progress":
            curr, total = args
            if total > 0: self.progress.set(curr / total)

    def _on_gemini_keys_exhausted(self, client): return False

    def _on_db_browser_click(self):
        from gui.db_browser_window import DbBrowserWindow
        from db.product_repo import ProductRepository
        DbBrowserWindow(self.master, self.db, ProductRepository(self.db), scheduler=self.scheduler)

    def _on_link_entry_click(self):
        from gui.direct_urls_window import DirectUrlsWindow
        def on_save(links):
            self.target_links = links
            self.msg_queue.put(("status_update", f"Targeting {len(links)} manual links."))
        DirectUrlsWindow(self.master, self.target_links, on_save=on_save)

    def _on_excel_click(self):
        from exporters.excel_exporter import ExcelExporter
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if file_path:
            ExcelExporter(file_path).export(self.all_found)
            messagebox.showinfo("Export", "Success")
