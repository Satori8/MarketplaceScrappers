"""
reports/snapshot_report.py
──────────────────────────────────────────────────────────────────────────
Generates a styled 4-sheet Excel report comparing 2+ snapshots (same task).

Sheets:
  1. Summary       – KPI block + 2 line charts
  2. Current Products – latest snapshot with price delta vs earliest
  3. Price Dynamics   – products present in 2+ snapshots, price changed
  4. Appeared / Disappeared

Usage:
    from reports.snapshot_report import generate_snapshot_report
    generate_snapshot_report([3, 7, 12], "report.xlsx", "marketplace_scraper.db")
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────── colour palette ──────────────────────────────

_DARK_BG     = "1E1E2E"
_WHITE       = "FFFFFF"
_LIGHT_GRAY  = "F9F9F9"
_GRAY_TEXT   = "888888"
_RED_FILL    = "FFDEDE"
_GREEN_FILL  = "DEFFDE"
_BLUE_FILL   = "DEF0FF"
_KPI_LABEL   = "AAAAAA"
_ACCENT      = "5B8DEF"
_THIN_BORDER_COLOR = "DDDDDD"

# ─────────────────────────────── helpers ─────────────────────────────────────

def _font(size=11, bold=False, color=None, name="Calibri") -> Font:
    kw = {"name": name, "size": size, "bold": bold}
    if color:
        kw["color"] = color
    return Font(**kw)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color=_THIN_BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _header_cell(ws, row, col, text, width_hint=None):
    """Dark header cell: white bold text on dark background."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = _font(13, bold=True, color=_WHITE)
    c.fill = _fill(_DARK_BG)
    c.alignment = _center()
    c.border = _thin_border()
    if width_hint and width_hint > (ws.column_dimensions[get_column_letter(col)].width or 0):
        ws.column_dimensions[get_column_letter(col)].width = width_hint
    return c


def _auto_width(ws, min_w=12, max_w=60):
    """Fit column widths to content."""
    for col_cells in ws.columns:
        best = 0
        for cell in col_cells:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                best = max(best, val_len)
            except Exception:
                pass
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(best + 2, max_w))


def _fmt_pct(val: float) -> str:
    return f"{val:+.1f}%"


def _pct(base, curr) -> Optional[float]:
    if base and base != 0:
        return (curr - base) / base * 100
    return None


# ─────────────────────────────── DB queries ──────────────────────────────────

def _connect(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _validate_same_task(conn: sqlite3.Connection, snapshot_ids: list[int]) -> int:
    """Return task_id if all snapshots belong to the same task, else raise."""
    placeholders = ",".join("?" * len(snapshot_ids))
    rows = conn.execute(
        f"SELECT DISTINCT task_id FROM snapshots WHERE id IN ({placeholders})",
        snapshot_ids,
    ).fetchall()
    if len(rows) != 1:
        raise ValueError(
            f"Snapshots belong to different tasks or some IDs are invalid: {snapshot_ids}"
        )
    return rows[0]["task_id"]


def _fetch_metadata(conn: sqlite3.Connection, task_id: int, snapshot_ids: list[int]):
    """Return (task_row, client_row, list_of_snapshot_rows ordered oldest→newest)."""
    task = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (task["client_id"],)).fetchone()
    placeholders = ",".join("?" * len(snapshot_ids))
    snapshots = conn.execute(
        f"SELECT * FROM snapshots WHERE id IN ({placeholders}) ORDER BY run_at ASC",
        snapshot_ids,
    ).fetchall()
    return task, client, snapshots


def _fetch_products_for_snapshot(conn: sqlite3.Connection, snapshot_id: int) -> list[dict]:
    """Return all products for a snapshot as dicts (URL is the identity key)."""
    rows = conn.execute(
        """
        SELECT
            sp.id          AS sp_id,
            sp.snapshot_id,
            COALESCE(p.url,  sp.url)           AS url,
            COALESCE(p.title, sp.name)          AS name,
            COALESCE(p.sku,  sp.sku)            AS sku,
            COALESCE(p.merchant_name, sp.merchant_name) AS merchant,
            sp.url_tag,
            sp.price,
            sp.avail_code
        FROM snapshot_products sp
        LEFT JOIN products p ON sp.product_id = p.id
        WHERE sp.snapshot_id = ?
        """,
        (snapshot_id,),
    ).fetchall()
    return [dict(r) for r in rows]


# ─────────────────────────────── Sheet 1 — Summary ───────────────────────────

def _build_summary_sheet(wb, task, client, snapshots, products_per_snap: dict):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    earliest_id = snapshots[0]["id"]
    latest_id   = snapshots[-1]["id"]
    earliest_products = products_per_snap[earliest_id]
    latest_products   = products_per_snap[latest_id]

    # Index by URL
    earliest_by_url = {p["url"]: p for p in earliest_products if p.get("url")}
    latest_by_url   = {p["url"]: p for p in latest_products if p.get("url")}

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total_latest = len(latest_products)
    new_count    = sum(1 for u in latest_by_url if u not in earliest_by_url)
    gone_count   = sum(1 for u in earliest_by_url if u not in latest_by_url)

    price_changed = 0
    max_pct_change = 0.0
    max_pct_name   = ""
    for url, lp in latest_by_url.items():
        ep = earliest_by_url.get(url)
        if ep and ep["price"] and lp["price"]:
            pct = _pct(ep["price"], lp["price"])
            if pct is not None and abs(pct) > 0.01:
                price_changed += 1
                if abs(pct) > abs(max_pct_change):
                    max_pct_change = pct
                    max_pct_name   = lp["name"] or url

    def avg_price(products):
        prices = [p["price"] for p in products if p.get("price")]
        return sum(prices) / len(prices) if prices else 0.0

    avg_earliest = avg_price(earliest_products)
    avg_latest   = avg_price(latest_products)

    date_first = (snapshots[0]["run_at"] or "")[:16].replace("T", " ")
    date_last  = (snapshots[-1]["run_at"] or "")[:16].replace("T", " ")
    generated  = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ── Header block (rows 1–5) ───────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "📊  MARKET SNAPSHOT REPORT"
    t.font  = _font(18, bold=True, color=_WHITE)
    t.fill  = _fill(_DARK_BG)
    t.alignment = _center()
    ws.row_dimensions[1].height = 38

    def info_row(row, label, value):
        ws.merge_cells(f"A{row}:C{row}")
        ws.merge_cells(f"D{row}:H{row}")
        a = ws[f"A{row}"]
        b = ws[f"D{row}"]
        a.value = label
        b.value = value
        a.font  = _font(11, bold=True, color=_GRAY_TEXT)
        b.font  = _font(11)
        a.fill  = _fill("1A1A2A")
        b.fill  = _fill("22223A")
        a.alignment = _left()
        b.alignment = _left()
        ws.row_dimensions[row].height = 22

    info_row(2, "  Client",    client["name"] if client else "—")
    info_row(3, "  Task",      task["title"] if task else "—")
    info_row(4, "  Period",    f"{date_first}  →  {date_last}")
    info_row(5, "  Generated", generated)

    ws.row_dimensions[6].height = 10  # spacer

    # ── KPI block (rows 7–11) ────────────────────────────────────────────────
    kpis = [
        ("Total Products\n(latest)", str(total_latest)),
        ("New Products",             str(new_count)),
        ("Disappeared",              str(gone_count)),
        ("Price Changes",            str(price_changed)),
        (f"Avg Price\nearliest → latest",
         f"₴{avg_earliest:,.0f} → ₴{avg_latest:,.0f}"),
        ("Max Price Change\n(product)",
         f"{_fmt_pct(max_pct_change)}\n{max_pct_name[:30]}"),
    ]

    # 6 KPIs across columns A-H (pairs of columns per KPI)
    kpi_cols = [1, 2, 3, 4, 5, 6, 7, 8]  # we'll use 2 cols per KPI in A-L
    ws.merge_cells("A7:L7")
    hdr = ws["A7"]
    hdr.value = "KEY PERFORMANCE INDICATORS"
    hdr.font  = _font(11, bold=True, color=_GRAY_TEXT)
    hdr.fill  = _fill("16162A")
    hdr.alignment = _center()
    ws.row_dimensions[7].height = 22

    for i, (label, value) in enumerate(kpis):
        col_start = i * 2 + 1   # 1, 3, 5, 7, 9, 11
        col_end   = col_start + 1

        ws.merge_cells(
            start_row=8, start_column=col_start,
            end_row=8, end_column=col_end
        )
        lc = ws.cell(row=8, column=col_start)
        lc.value = label
        lc.font  = _font(9, color=_KPI_LABEL)
        lc.fill  = _fill("1A1A2E")
        lc.alignment = _center(wrap=True)
        ws.row_dimensions[8].height = 30

        ws.merge_cells(
            start_row=9, start_column=col_start,
            end_row=9, end_column=col_end
        )
        vc = ws.cell(row=9, column=col_start)
        vc.value = value
        vc.font  = _font(14, bold=True, color=_WHITE)
        vc.fill  = _fill(_DARK_BG)
        vc.alignment = _center(wrap=True)
        ws.row_dimensions[9].height = 40

    ws.row_dimensions[10].height = 14  # spacer

    # ── Chart data table (hidden rows, used as chart source) ─────────────────
    chart_start_row = 12
    ws.cell(row=chart_start_row, column=1, value="Snapshot").font = _font(9, bold=True)
    ws.cell(row=chart_start_row, column=2, value="Product Count").font = _font(9, bold=True)
    ws.cell(row=chart_start_row, column=3, value="Avg Price (₴)").font = _font(9, bold=True)

    for i, snap in enumerate(snapshots):
        r = chart_start_row + 1 + i
        label = (snap["run_at"] or f"#{snap['id']}")[:16].replace("T", " ")
        prods = products_per_snap[snap["id"]]
        prices = [p["price"] for p in prods if p.get("price")]
        avg    = sum(prices) / len(prices) if prices else 0.0
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=len(prods))
        ws.cell(row=r, column=3, value=round(avg, 2))

    n_snaps = len(snapshots)
    data_rows = range(chart_start_row + 1, chart_start_row + 1 + n_snaps)
    last_data_row = chart_start_row + n_snaps

    # ── Chart 1: product count ────────────────────────────────────────────────
    chart1 = LineChart()
    chart1.title  = "Product Count per Snapshot"
    chart1.style  = 10
    chart1.y_axis.title = "Products"
    chart1.x_axis.title = "Snapshot"
    chart1.width  = 20
    chart1.height = 12

    data_ref1 = Reference(ws, min_col=2, max_col=2,
                          min_row=chart_start_row, max_row=last_data_row)
    cats1 = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=last_data_row)
    chart1.add_data(data_ref1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws.add_chart(chart1, "A12")

    # ── Chart 2: average price ────────────────────────────────────────────────
    chart2 = LineChart()
    chart2.title  = "Avg Price per Snapshot (₴)"
    chart2.style  = 10
    chart2.y_axis.title = "Price ₴"
    chart2.x_axis.title = "Snapshot"
    chart2.width  = 20
    chart2.height = 12

    data_ref2 = Reference(ws, min_col=3, max_col=3,
                          min_row=chart_start_row, max_row=last_data_row)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats1)
    ws.add_chart(chart2, "K12")

    # ── Column sizing ─────────────────────────────────────────────────────────
    for i in range(1, 13):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "A2"


# ─────────────────────────── Sheet 2 — Current Products ──────────────────────

def _build_current_sheet(wb, snapshots, products_per_snap: dict):
    ws = wb.create_sheet("Current Products")
    ws.sheet_view.showGridLines = False

    earliest = snapshots[0]
    latest   = snapshots[-1]
    earliest_by_url = {p["url"]: p for p in products_per_snap[earliest["id"]] if p.get("url")}
    latest_products = products_per_snap[latest["id"]]

    COLS = ["Name", "SKU", "Merchant", "Tag", "Price (latest)", "Price (earliest)", "Delta ₴", "Delta %", "URL"]
    for ci, h in enumerate(COLS, 1):
        _header_cell(ws, 1, ci, h)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes = "A2"

    fill_red   = _fill(_RED_FILL)
    fill_green = _fill(_GREEN_FILL)
    fill_blue  = _fill(_BLUE_FILL)
    fill_alt   = _fill(_LIGHT_GRAY)
    fill_white = _fill(_WHITE)

    for ri, p in enumerate(latest_products, 2):
        url         = p.get("url") or ""
        ep          = earliest_by_url.get(url)
        price_now   = p.get("price") or 0.0
        price_then  = ep["price"] if ep and ep.get("price") else None
        is_new      = ep is None

        if price_then is not None and price_then != 0:
            delta_abs = price_now - price_then
            delta_pct = _pct(price_then, price_now)
        else:
            delta_abs = None
            delta_pct = None

        row_vals = [
            p.get("name") or "",
            p.get("sku") or "",
            p.get("merchant") or "",
            p.get("url_tag") or "",
            price_now,
            price_then if price_then is not None else "",
            round(delta_abs, 2) if delta_abs is not None else "",
            _fmt_pct(delta_pct) if delta_pct is not None else "",
            url,
        ]
        ws.append(row_vals)

        # Row fill
        if is_new:
            row_fill = fill_blue
        elif delta_pct and delta_pct > 0.0:
            row_fill = fill_red
        elif delta_pct and delta_pct < 0.0:
            row_fill = fill_green
        else:
            row_fill = fill_alt if ri % 2 == 0 else fill_white

        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.fill = row_fill
            cell.font = _font(10)
            cell.alignment = _left()

    _auto_width(ws)


# ─────────────────────────── Sheet 3 — Price Dynamics ────────────────────────

def _build_price_dynamics_sheet(wb, snapshots, products_per_snap: dict):
    ws = wb.create_sheet("Price Dynamics")
    ws.sheet_view.showGridLines = False

    snap_labels = [(s["id"], (s["run_at"] or f"snap#{s['id']}")[:16].replace("T", " "))
                   for s in snapshots]

    # Find products present in 2+ snapshots AND with at least one price change
    url_to_prices: dict[str, dict[int, float]] = {}
    url_to_name: dict[str, str] = {}
    url_to_tag: dict[str, str] = {}

    for snap in snapshots:
        for p in products_per_snap[snap["id"]]:
            u = p.get("url")
            if not u:
                continue
            url_to_prices.setdefault(u, {})[snap["id"]] = p.get("price") or 0.0
            if u not in url_to_name:
                url_to_name[u] = p.get("name") or u
            if u not in url_to_tag and p.get("url_tag"):
                url_to_tag[u] = p.get("url_tag")

    # Filter: present in 2+ AND has a price change
    qualifying = []
    for url, price_map in url_to_prices.items():
        if len(price_map) < 2:
            continue
        prices = list(price_map.values())
        if len(set(prices)) > 1:  # at least one price change
            qualifying.append(url)

    # Build header
    header = ["Name", "Tag", "URL"] + [lbl for _, lbl in snap_labels]
    for ci, h in enumerate(header, 1):
        _header_cell(ws, 1, ci, h)
    ws.freeze_panes = "A2"

    fill_red   = _fill(_RED_FILL)
    fill_green = _fill(_GREEN_FILL)
    fill_gray  = _fill("EEEEEE")
    fill_alt   = _fill(_LIGHT_GRAY)
    fill_white = _fill(_WHITE)

    for ri, url in enumerate(qualifying, 2):
        price_map = url_to_prices[url]
        row_base = [url_to_name.get(url, ""), url_to_tag.get(url, ""), url]
        ws.append(row_base + [""] * len(snap_labels))

        base_fill = fill_alt if ri % 2 == 0 else fill_white

        # Name, Tag, URL cols
        for ci in (1, 2, 3):
            cell = ws.cell(row=ri, column=ci)
            cell.value = row_base[ci - 1]
            cell.font  = _font(10)
            cell.alignment = _left()
            cell.fill = base_fill

        prev_price = None
        for col_offset, (snap_id, _) in enumerate(snap_labels):
            ci = 4 + col_offset
            cell = ws.cell(row=ri, column=ci)
            price = price_map.get(snap_id)
            if price is None:
                cell.fill = fill_gray
                cell.value = ""
            else:
                cell.value = price
                cell.font  = _font(10, bold=True)
                cell.alignment = _center()
                if prev_price is None:
                    cell.fill = base_fill
                elif price > prev_price:
                    cell.fill = fill_red
                elif price < prev_price:
                    cell.fill = fill_green
                else:
                    cell.fill = base_fill
                prev_price = price

    _auto_width(ws)


# ──────────────────────── Sheet 4 — Appeared / Disappeared ───────────────────

def _build_appeared_sheet(wb, snapshots, products_per_snap: dict):
    ws = wb.create_sheet("Appeared - Disappeared")
    ws.sheet_view.showGridLines = False

    earliest = snapshots[0]
    latest   = snapshots[-1]

    early_by_url  = {p["url"]: p for p in products_per_snap[earliest["id"]] if p.get("url")}
    latest_by_url = {p["url"]: p for p in products_per_snap[latest["id"]] if p.get("url")}

    # When was a URL first seen across ALL snapshots
    first_seen: dict[str, str] = {}
    last_seen: dict[str, str] = {}
    for snap in snapshots:
        label = (snap["run_at"] or f"#{snap['id']}")[:16].replace("T", " ")
        for p in products_per_snap[snap["id"]]:
            u = p.get("url")
            if not u:
                continue
            if u not in first_seen:
                first_seen[u] = label
            last_seen[u] = label

    new_products  = [p for u, p in latest_by_url.items() if u not in early_by_url]
    gone_products = [p for u, p in early_by_url.items() if u not in latest_by_url]

    fill_blue  = _fill(_BLUE_FILL)
    fill_red   = _fill(_RED_FILL)
    fill_alt   = _fill(_LIGHT_GRAY)
    fill_white = _fill(_WHITE)

    row = 1

    # ── Section 1: New Products ───────────────────────────────────────────────
    COLS_NEW = ["Name", "SKU", "Merchant", "Tag", "Price", "URL", "First Seen"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS_NEW))
    sec_hdr = ws.cell(row=row, column=1, value="🆕  NEW PRODUCTS")
    sec_hdr.font      = _font(13, bold=True, color=_WHITE)
    sec_hdr.fill      = _fill("1A3A5C")
    sec_hdr.alignment = _center()
    ws.row_dimensions[row].height = 26
    row += 1

    for ci, h in enumerate(COLS_NEW, 1):
        _header_cell(ws, row, ci, h)
    row += 1

    for i, p in enumerate(new_products):
        u = p.get("url") or ""
        vals = [
            p.get("name") or "",
            p.get("sku") or "",
            p.get("merchant") or "",
            p.get("url_tag") or "",
            p.get("price") or "",
            u,
            first_seen.get(u, ""),
        ]
        ws.append(vals)
        rf = fill_alt if i % 2 == 0 else fill_white
        for ci in range(1, len(COLS_NEW) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.fill = rf
            cell.font = _font(10)
            cell.alignment = _left()
        row += 1

    # Spacer
    row += 1
    ws.row_dimensions[row].height = 10
    row += 1

    # ── Section 2: Disappeared Products ──────────────────────────────────────
    COLS_GONE = ["Name", "SKU", "Merchant", "Tag", "Last Price", "URL", "Last Seen"]
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS_GONE))
    sec_hdr2 = ws.cell(row=row, column=1, value="💨  DISAPPEARED PRODUCTS")
    sec_hdr2.font      = _font(13, bold=True, color=_WHITE)
    sec_hdr2.fill      = _fill("3A1A1A")
    sec_hdr2.alignment = _center()
    ws.row_dimensions[row].height = 26
    row += 1

    for ci, h in enumerate(COLS_GONE, 1):
        _header_cell(ws, row, ci, h)
    row += 1

    for i, p in enumerate(gone_products):
        u = p.get("url") or ""
        vals = [
            p.get("name") or "",
            p.get("sku") or "",
            p.get("merchant") or "",
            p.get("url_tag") or "",
            p.get("price") or "",
            u,
            last_seen.get(u, ""),
        ]
        ws.append(vals)
        rf = fill_alt if i % 2 == 0 else fill_white
        for ci in range(1, len(COLS_GONE) + 1):
            cell = ws.cell(row=row, column=ci)
            cell.fill = rf
            cell.font = _font(10)
            cell.alignment = _left()
        row += 1

    ws.freeze_panes = "A2"
    _auto_width(ws)


# ─────────────────────────── Public API ──────────────────────────────────────

def generate_snapshot_report(
    snapshot_ids: list[int],
    output_path: str,
    db_path: str,
) -> None:
    """
    Generate a 4-sheet Excel report comparing 2+ snapshots from the same task.

    Parameters
    ----------
    snapshot_ids : list[int]
        Ordered oldest → newest. Must be from the same task. Min length: 2.
    output_path  : str
        Full path to the output .xlsx file.
    db_path      : str
        Path to the SQLite database file.

    Raises
    ------
    ValueError
        If fewer than 2 snapshots are provided or they belong to different tasks.
    """
    if len(snapshot_ids) < 2:
        raise ValueError(f"At least 2 snapshot IDs are required, got {len(snapshot_ids)}.")

    conn = _connect(db_path)
    try:
        task_id = _validate_same_task(conn, snapshot_ids)
        task, client, snapshots = _fetch_metadata(conn, task_id, snapshot_ids)

        if len(snapshots) < 2:
            raise ValueError("Could not find 2+ valid snapshots for the given IDs.")

        # Pre-fetch all products once per snapshot
        products_per_snap: dict[int, list[dict]] = {}
        for snap in snapshots:
            products_per_snap[snap["id"]] = _fetch_products_for_snapshot(conn, snap["id"])

    finally:
        conn.close()

    # Build workbook
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, task, client, snapshots, products_per_snap)
    _build_current_sheet(wb, snapshots, products_per_snap)
    _build_price_dynamics_sheet(wb, snapshots, products_per_snap)
    _build_appeared_sheet(wb, snapshots, products_per_snap)

    wb.save(output_path)
